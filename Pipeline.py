# python -m spacy download en_core_web_sm

import pandas as pd
import io
import librosa
import numpy as np
import openai
import requests
from time import sleep
import whisper
import os.path
from pathlib import Path 
import openpyxl
from os.path import exists
from transformers import GPT2Tokenizer
from openpyxl import load_workbook
# import seaborn as sns
import spacy
from textblob import TextBlob
# import matplotlib.pyplot as plt
import re

# set audio path of audio file
# audio_path = "f7422f49-5b42-48aa-a98c-67f1da3bda21_0_r.wav"

# with open("audio_file.wav", "wb") as f:
# audio_path = r"audio_file.wav"
    # audio_path = r"c0fec98f-91de-4349-b21d-25dc05f55eaf_0_r.wav"

# if __name__ == "__main__":
    # code to be executed if this file is run as the main program


class Audio_to_text:
    """
    A class that transcribes audio files into text and provides a prompt-based evaluation of the transcription.
    
    Attributes:
    audio_file (str): the name of the audio file to transcribe
    saved_response (str): the saved transcription of the audio file
    audio_df (pandas.DataFrame): a DataFrame that stores the audio file names and their corresponding transcriptions
    
    Methods:
    load_audio_transcribe(): loads the audio file names and their corresponding transcriptions from a CSV file
    fetch_audio(): transcribes the audio file and saves the transcription in the audio_df DataFrame
    write_df_toexcel(output_df, file_name, sheet_name): writes a pandas DataFrame to an Excel file
    response_prompt(model, temperature, max_tokens, top_p, frequency_penalty, presence_penalty):
        prompts the user to provide scores for a set of predefined prompts based on the audio transcription
    """
    def __init__(self, audio_file):
        # Initialising
        self.audio_file = audio_file
        self.saved_response = ""
        self.audio_df = None
    
    def load_audio_transcribe(self):
        """Summary: Create DataFrame for Audio Transcribe if not exists
        """
        file_exists = exists('Audio_Transcription.xlsx')
        if not file_exists:
            self.audio_df = pd.DataFrame(columns=["file_name", "transcribe"])
            self.audio_df.to_excel("Audio_Transcription.xlsx", index=False)
        else:
            self.audio_df = pd.read_excel("Audio_Transcription.xlsx")

    def fetch_audio(self):
        """Summary: This function will fetch the audio and will convert audio file to text file.
            Return: Str(It will return audio text)
        """
        # Read Transcribe csv file if extry exists for the audio file    
        self.load_audio_transcribe()
        
        if self.audio_file in self.audio_df["file_name"].values:
            self.saved_response = self.audio_df[self.audio_df["file_name"]==self.audio_file]["transcribe"].values[0]
        else:
            model = whisper.load_model("base")
            print("self audio:",self.audio_file)
            
            # assume self.audio_file is an UploadedFile object
            audio_bytes = self.audio_file.read()
            audio_np, sr = librosa.load(io.BytesIO(audio_bytes))
            result = model.transcribe(audio_np)
            self.saved_response = result['text']
            self.audio_df = self.audio_df.append({"file_name":self.audio_file.name, "transcribe":self.saved_response}, ignore_index = True)
            self.af, self.sr= self.saved_response, sr
            self.audio_df.to_excel("Audio_Transcription.xlsx", index=False)
            
            # get the path of the uploaded file
            audio_file_path = Path(self.audio_file.name)
            
            # call assign_score with the path of the file
            # score_obj.assign_score(df, audio_file_path.stem + '_1.xlsx')

#         print(self.saved_response)
    
    def write_df_toexcel(self, output_df, file_name, sheet_name):
        # Try to open an existing Excel file and load the sheet
        try:
            book = load_workbook(file_name)
            writer = pd.ExcelWriter(file_name, engine='openpyxl') 
            writer.book = book
            # Check if the sheet name already exists
            if sheet_name in book.sheetnames:
                # Remove the existing sheet
                book.remove(book[sheet_name])
            # Write the DataFrame to the sheet
#             output_df.to_excel(writer, sheet_name=sheet_name, header=False)
            # Save the workbook
#             writer.save()
            print(f'Successfully updated sheet {sheet_name} in {file_name}.')
        except FileNotFoundError:
            output_df.to_excel(file_name, sheet_name=sheet_name, header=False)
            print(f'Successfully created sheet {sheet_name} in {file_name}.')

    def response_prompt(self, model="text-davinci-003", temperature=0.7, length_of_prompt=1, max_tokens=256, top_p=1, frequency_penalty=0, presence_penalty=0):
        '''Summary: This function will provide the scores of the each parameters
            Args: list of prompts which we are getting from generate_text function 
                and also the default values have been set for OpenAI response parameters
            Return: DataFrame(It will return DataFrame with score and Parameter columns)'''
        
        # length_of_prompt=int(length_of_prompt)

        # Getting Audio Transcribe
        self.fetch_audio()

        # Parameter list for DataFrame
        df_Parameter_list = ['ISR Introduction',
                            'Parent/Student Details Confirmation',
                            'Update on qualification to next round',
                            'Explaining the report positively followed by area of improvement',
                            "Explain the scholarship awarded",
                            'Program Structure-3 modules-6 months-start with a module',
                            'Value Vs Price - reaffirm on #sessions/duration followed by module cost',
                            'Explained about batch details',
                            'Active listening/Understanding/Comprehending/stammering/incorrect sentence',
                            'IST informed its not program fee its registration fee',
                            'IF Hots Program is not taken then you are not allowed for the next level of exam',
                            'Complete 3 modules for 3000 or lesser price',
                            'informed 1Lac or 50k cash prize or Govt scholar',
                            'This is one time investment no further payments in the future',
                            'Malpractice to increase Productivity/Metrics',
                            'Rude/argumentative/Abusive on call/Mocking/Sarcastic']

        # List of prompts to check different parameters
        param_list = ['0 to 3, if the caller introduced themselves',
                    '0 to 3, if the caller confirmed about Parents and Student Details',
                    '0 to 3, if the caller informed parents that his or her child has qualified for the next round',
                    '0 to 3, if the caller explained the report positively followed by the area of improvement',
                    '0 to 3, if the caller explained the scholarship awarded',
                    '0 to 3, if the caller explained the Program Structure-3 modules-6 months-start with a module',
                    '0 to 3, if the caller reaffirmed on Value Vs Price - #sessions/duration followed by module cost',
                    '0 to 3, if the caller explained about batch details',
                    '0 to 3, if the caller demonstrated active listening/Understanding/Comprehending/stammering/incorrect sentence',
                    '0 to 3, if the caller informed that IST is not a program fee, but a registration fee',
                    '0 to 3, if the caller informed that if HOTS Program is not taken, then they are not allowed for the next level of exam',
                    '0 to 3, if the caller informed that they can complete 3 modules for 3000 or lesser price',
                    '0 to 3, if the caller informed that there is 1Lac or 50k cash prize or Govt scholar',
                    '0 to 3, if the caller informed that this is a one-time investment and there are no further payments in the future',
                    '0 to 3, if the caller explained the Malpractice to increase Productivity/Metrics',
                    '0 to 3, if the caller was Rude/argumentative/Abusive on call/Mocking/Sarcastic'] 

        custom_df_type_list=[]
        df_score_list = []
        call_score = pd.DataFrame()
        
        #To provide excel name         
        # excel_name=Path(audio_path).stem+ '.xlsx'
        excel_name = Path(self.audio_file.name).stem + '.xlsx'

        print(excel_name)
        
        with open("Api_key.txt", "r") as f:
            api_key=f.read()   
        
        # Getting openai response  
        for i in range(length_of_prompt):
            print(param_list[i])
            custom_df_type_list.append(df_Parameter_list[i])
            prmt_param = self.saved_response + " Provide the score only between " + param_list[i]
        
            openai.api_key = api_key
            response = openai.Completion.create(
            model=model,
            prompt=prmt_param,
            temperature=temperature,
            max_tokens=max_tokens,
            top_p=top_p,
            frequency_penalty=frequency_penalty,
            presence_penalty=presence_penalty
            )

            if response != '':
                prompt_response = response['choices'][0]['text']
                df_score_list.append(prompt_response)
            else:
                print("Response is empty")    
                
        
        # appending the list values to columns 
        df2 = pd.DataFrame({'file_name': [self.audio_file.name], 'transcribe': [self.saved_response]})
        score_df = pd.DataFrame(columns=["Score"])

        score_df["Score"] = df_score_list
        score_df['Score']= score_df.Score.str.extract('(\d+)')
        score_list = score_df['Score'].values
        dict_call = {}
        for key in custom_df_type_list:
            for value in score_list:
                dict_call[key] = value
                break
        # Printing resultant dictionary
        data = [dict_call]

        # df = pd.DataFrame((res))
        df1=pd.DataFrame.from_dict(data)
        print(df1)
        open_Ai_Score=df2.join(df1)

    # self.write_df_toexcel(open_Ai_Score, file_name=excel_name, sheet_name=f"{model}-{temperature}" )

        return open_Ai_Score

# obj = Audio_to_text(audio_file = audio_path)
# obj.response_prompt(temperature=0.7) 
    
# ----------------------------------------------keyword extraction and subjective parmeters scores-----------------------------------------------------
class VoiceParameterEvaluator:
    
    def proper_greet(self, dataframe):
        """
        This function takes in a pandas dataframe containi \ng the transcribed sales call data and returns a modified dataframe 
        with a score and reason column based on the presence of a proper greeting in the call.

        Parameters
        dataframe : A DataFrame containing the transcribed sales call data.

        Returns
        A modified DataFrame with a 'greet_score' and 'greet_reason' column added.
        """
        pattern = r'(hello|hi|hey|very good morning|good morning|very good afternoon|good afternoon|very good evening|good evening),?\s+(mam|madam|sir)[.!?]?'

        # Check if the call starts with a valid greeting
        valid_greeting = dataframe['transcribe'].str.lower().str.extract(pattern, flags=re.IGNORECASE)

        dataframe['valid_greeting'] = valid_greeting.iloc[:, 1].notna()

        # Add score column as integer indicating quality of call
        dataframe['Call opening with the right greeting'] = dataframe['valid_greeting'].astype(int) * 3

        # Provide reason for score assigned
        dataframe['greet_reason'] = dataframe.apply(lambda row: 'Because caller greeted properly' if row['valid_greeting'] else 'Because caller did not greet properly', axis=1)

        return dataframe

    def what_hots(self, pitch):
        """
        This function returns a tuple containing a score and reason based on whether the caller mentioned the HOTS program 
        and specific skills that can be improved.

        Parameters
        pitch : str
            The transcribed text of the sales call.

        Returns
        tuple
            A tuple containing a score (0/1/3) and reason for the score.
        """
        hots_keywords = ['hots program','higher order thinking skills','hots']
        skills_keywords = ['iq','logical thinking','cognitive skills','aptitude', 'smart thinking','visual analogy',
                        'compare  contrast','cause and effect', 'main idea stated','main idea unstated', 'sequencing', 
                        'verbal anlogy', 'making conclusion', 'classification', 'algorithmic thinking','problem solver',
                        'problem solving', 'decision making','life skills']
        
        pitch = pitch.lower()
        found_hots_keywords = [keyword for keyword in hots_keywords if keyword.lower() in pitch]
        found_skills_keywords = [keyword for keyword in skills_keywords if keyword.lower() in pitch]
        if found_hots_keywords:
            if len(found_skills_keywords) >= 2:
                return (3, f"Caller mentioned {', '.join(found_hots_keywords)} and at least 2 specific skills that can be improved: {', '.join(found_skills_keywords)}")
            elif len(found_skills_keywords) == 1:
                return (1, f"Caller mentioned {', '.join(found_hots_keywords)} but only mentioned 1 specific skill that can be improved: {', '.join(found_skills_keywords)}")
            else:
                return (0, f"Caller mentioned {', '.join(found_hots_keywords)} but did not mention any specific skills")
        else:
            return (0, "Caller did not mention HOTS program")

    # Define function to get score for explaining 3 skills parameter
    def explain_skills(self, row):
        """
        This function takes in a pandas dataframe row containing the transcribed sales call data and returns a dictionary 
        with score and reason for skills mentioned in the call
        """

        # Define the skills and their respective keywords
        skills = ['logical thinking', 'critical thinking', 'problem solving']
        logical_thinking =['reasoning based on facts', 'reasoning',
                'identifying patterns', 'drawing conclusions', 'logic and evidence']
        critical_thinking = ['analyzing information', 'evaluating information', 'ability to understand',
                'making judgments', 'making decisions']
        problem_solving =['identifying problem', 'defining the problem', 'analyzing a situation',
                'ability to understand', 'ability to solve problems', 'form decision']   

    # Initialize the score and reason variables
        score = 0
        reason = ''

        # Convert the 'transcribe' column to lowercase
        transcribe = row['transcribe'].lower()

        # Check if any of the keywords in the 'skills' list are present in the 'transcribe' column
        if any(skill in transcribe for skill in skills):
            # Check for keywords in the 'logical_thinking' list
            found_keywords_logical = [keyword for keyword in logical_thinking if keyword in transcribe]
            if found_keywords_logical:
                score += 1
                reason += 'Found keywords in logical_thinking: ' + ', '.join(found_keywords_logical) + '\n'

            # Check for keywords in the 'critical_thinking' list
            found_keywords_critical = [keyword for keyword in critical_thinking if keyword in transcribe]
            if found_keywords_critical:
                score += 1
                reason += 'Found keywords in critical_thinking: ' + ', '.join(found_keywords_critical) + '\n'

            # Check for keywords in the 'problem_solving' list
            found_keywords_solving = [keyword for keyword in problem_solving if keyword in transcribe]
            if found_keywords_solving:
                score += 1
                reason += 'Found keywords in problem_solving: ' + ', '.join(found_keywords_solving) + '\n'

            # Assign score 1 if no keywords found in any of the 3 lists
            if score == 0:
                score = 1
                reason = 'No keywords found in logical_thinking, critical_thinking, or problem_solving'

            # Assign score 3 if keywords found in any 2 lists out of the remaining 3 lists
            elif score == 1 or score == 2:
                if found_keywords_logical and found_keywords_critical:
                    score += 1
                    reason += 'Found keywords in logical_thinking and critical_thinking\n'
                elif found_keywords_logical and found_keywords_solving:
                    score += 1
                    reason += 'Found keywords in logical_thinking and problem_solving\n'
                elif found_keywords_critical and found_keywords_solving:
                    score += 1
                    reason += 'Found keywords in critical_thinking and problem_solving\n'

        else:
            score = 0
            reason = 'No keywords found in skills list.'

        # Return the score and reason as a tuple
        return score, reason
    

    def why_hots(self, text):
        """
        Calculates scores and reasons based on the number of benefits explained in the given text.

        Args:
        text (str): Text to be evaluated.

        Returns:
        tuple(int, str): Score and reason.
        """
        # Define keyword lists
        memorisation = ['not memorize', 'not by memorizing', 'rather than memorizing']
        benefits = ['concentration', 'understand the concept','success in academics', 'competitive exams', 'career', 'jobs']
        progress_reports =['Monthly Progress Reports','10-15 mins of in app practices', 'Weekly Quiz']

        # Initialize spaCy with the English model
        nlp = spacy.load('en_core_web_sm')

        score = 0
        found_keywords = []

        # Lemmatize the text using spaCy
        doc = nlp(text)
        lemmatized_text = " ".join([token.lemma_ for token in doc])

        # Convert lemmatized text to lowercase
        lemmatized_text = lemmatized_text.lower()

        # Check if lemmatized text contains keywords from each list
        if any(keyword in lemmatized_text for keyword in memorisation):
            found_keywords.append(next(keyword for keyword in memorisation if keyword in lemmatized_text))
        if any(keyword in lemmatized_text for keyword in benefits):
            found_keywords.append(next(keyword for keyword in benefits if keyword in lemmatized_text))
        if any(keyword in lemmatized_text for keyword in progress_reports):
            found_keywords.append(next(keyword for keyword in progress_reports if keyword in lemmatized_text))

        # Determine score and reason based on number of found keywords
        if len(found_keywords) >= 2:
            score = 3
            reason = ", ".join(found_keywords)
        elif len(found_keywords) == 1:
            score = 1
            reason = found_keywords[0]
        else:
            reason = "No relevant keywords found"

        return score, reason

    
    def teach_method(self, row):        
        """
        Calculates the score and reason for a given row based on the keywords found in the 'transcribe' column.

        Parameters:
        row (pandas.Series): A pandas Series containing the transcription and other details for a given row

        Returns:
        pandas.Series: A pandas Series containing the score and reason for the given row

        """
        # Define keyword lists
        weekly = ['two classes per week', 'two classes', 'two weekly classes', 'weekly 2 classes','twice a week',
                    '2 class','weekly twice']
        about_class = ['interactive session', 'lmt','same level of understanding','four to five students', 
                    'level mapping test','4 students','4 to 6 students','interactive classes']                   
        after_class = ['10 worksheets', 'projects', 'skill card', 'progress reports',]

        score = 0
        reason = ''
        for kp in weekly:
            if kp in row['transcribe'].lower():
                score += 1
                reason += f"{kp}, "
                # Break after first match if the keyword is in the 'two_classes' list
                break

        for kp in about_class:
            if kp in row['transcribe']:
                score += 1
                reason += f"{kp}, "

        for kp in after_class:
            if kp in row['transcribe']:
                score += 1
                reason += f"{kp}, "

        # Remove the last comma and space from the reason string
        reason = reason[:-2]

        if score == 0:
            reason = 'No keypoints covered'
        elif score == 1:
            reason = 'One keypoint covered:\n' + reason
        else:
            reason = f'{score} keypoints covered:\n' + reason
            if score >= 2:
                score = 3

        return pd.Series({'score': score, 'reason': reason})      
    
    def batch_details(self, row):
        """
        This function scores the batch details mentioned in the given row of the input dataframe.
        It looks for specific keywords in the 'transcribe' column and returns a score and reason based on the presence of those keywords.

        Parameters:
        row (pandas.core.series.Series): A single row of input data to score.

        Returns:
        pandas.core.series.Series: A pandas Series containing the score and reason for the batch details scored in the given row.
        """
        # Define keyword lists
        online_classes = ['online class', 'google meet']
        links = ['links', 'login credentials', 'shared with students']
        recording = ['recorded classes','missed live sessions']
        support = ['counselor', 'certified teacher','mentor','faculty','instructor']
        track = ['child\'s progress','student progress','track progress']
        time = ['one hour','45 minutes','1 hour']
        keywords = [online_classes, links, recording, support, track, time]
        found = []

        for k in keywords:
            found.extend([x for x in k if x in row['transcribe'].lower()])

        if len(found) == 0:
            score = 0
            reason = 'No keywords found'
        elif len(found) == 1:
            score = 1
            reason = f'Found keyword: {found[0]}'
        else:
            score = 3
            reason = f'Found keywords: {", ".join(found)}'

        return pd.Series({'score': score, 'reason': reason})

    def wrong_price(self, row):        
        """
        Score row based on presence of keywords related to price, discounts, fees and taxes in the transcribed text
        Returns a score and reason for the score
        """
        # Define keyword lists
        price =['amount', 'price', 'enrollment','charges','registration','fee']
        course_fee = ['5000 rupees','rupees 5500']
        discount = ['extra discount','discount', '1000 rupees','rupees 1000']
        final_amount = ['rupees 2000', '2000 rupees','1500 rupees']
        gst = ['gst', 'including gst']
        
        transcribe = row['transcribe'].lower()
        keywords_found = []
        for keyword_list in [price, course_fee, discount, final_amount, gst]:
            found_keywords = [kw for kw in keyword_list if kw in transcribe]
            if len(found_keywords) >= 2:
                keywords_found.extend(found_keywords)
                break
            elif len(found_keywords) == 1:
                keywords_found.extend(found_keywords)

        if len(keywords_found) >= 2:
            reason = f"Keywords found: {', '.join(list(set(keywords_found)))}"
            return 3, reason
        elif len(keywords_found) == 1:
            reason = f"Keyword found: {', '.join(list(set(keywords_found)))}"
            return 1, reason
        else:
            return 0, 'No keywords found'

    # Define a function to calculate polarity and subjectivity scores
    def get_sentiment_scores(self,text):
        blob = TextBlob(text)
        return blob.sentiment.polarity, blob.sentiment.subjectivity

    # Define functions to score each of the four subjective parameters
    def score_rapport_building(self,row):
        if row['polarity_score'] > 0.15:
            return 3, 'Polarity score is high'
        else:
            return 0, 'Polarity score is low'

    def score_rigor_for_sale(self,row):
        if row['polarity_score'] > 0.2 and row['subjectivity_score'] > 0.2:
            return 0, 'Both scores are high'
        else:
            return 3, 'Both scores are not high'

    def score_enthusiastic(self,row):
        if row['subjectivity_score'] > 0.2:
            return 3, 'Subjectivity score is high'
        else:
            return 0, 'Subjectivity score is low'

    def score_objections_handled(self,row):
        if row['subjectivity_score'] > 0.15:
            return 3, 'Subjectivity score is high'
        else:
            return 0, 'Subjectivity score is low'            
        
    # Define function to score each row
    def assign_score(self,input_df,output_file):
        """
        This function takes an input Excel file path containing a transcribe column, and scores each row based on certain 
        keywords related to several sales parameters. The function returns a modified pandas DataFrame with the score and 
        reason of that score for each parameter. The function then saves the modified DataFrame to the output file in Excel format.

        :param input_file: The path to the input Excel file.
        :type input_file: str
        :param output_file: The path to the output Excel file.
        :type output_file: str
        :return: None
        """

        # Score greeting and add results to new column
        df = self.proper_greet(input_df)

        # Score pitch and add results to new column
        df['What is HOTS Program - improve IQ/Aptitude/Smart Thinking / Problem Solver/Decision Making'], df['hots_reason'] = zip(*df['transcribe'].apply(self.what_hots))

        # Score skills and add results to new columns
        df[["Explained Logical Thinking / Critical Thinking / Problem solving", "skill_reason"]] = df.apply(self.explain_skills, axis=1, result_type="expand")

        # Apply function to dataframe
        df[['Why HOTS: Minimum 2 benefits to be explained (Not to Memorisation/Concentration/Competitive Exams/Jobs)', 'why_hotreason']] = df['transcribe'].apply(self.why_hots).apply(pd.Series)

        # Apply function to each row
        df[['Teaching method - 2 class per week/ understanding skills & practicing the skills / LMT / Projects / Skill Cards / progress Report', 't_method_reason']] = df.apply(self.teach_method, axis=1, result_type="expand") 

        # Apply function to each row
        df[['Explained about batch details', 'batch_reason']] = df.apply(self.batch_details, axis=1, result_type="expand")

        # Apply the get_score function to each row of the dataframe
        df[['Incorrect Pricing / GST not informed (IST informed wrong price less than 2500rs)','price_reason']] = df.apply(self.wrong_price, axis=1, result_type="expand")

        # Apply the function to the transcribe column and create new columns for the scores
        df[['polarity_score', 'subjectivity_score']] = df['transcribe'].apply(self.get_sentiment_scores).apply(pd.Series)

        # Apply the functions to the data and create new columns for each parameter
        df[['Rapport Building', 'rapport_building_reason']] = df.apply(self.score_rapport_building, axis=1).apply(pd.Series)
        df[['Rigour for sale (Push for sale)', 'rigor_for_sale_reason']] = df.apply(self.score_rigor_for_sale, axis=1).apply(pd.Series)
        df[['Enthusiastic/Energy', 'enthusiastic_reason']] = df.apply(self.score_enthusiastic, axis=1).apply(pd.Series)
        df[['Was all objections handled well with all right rebuttals', 'objections_handled_reason']] = df.apply(self.score_objections_handled, axis=1).apply(pd.Series)

        # Remove column name 'valid_greeting' we just used to give score
        df = df.drop(['valid_greeting'], axis=1)
        # df = df.drop([['polarity_score', 'subjectivity_score']], axis=1)
        df = df.reset_index(drop=True)
        df.to_excel(output_file, index=False)
        df


# # create an instance of the Audio_to_text class and get the output DataFrame
# obj = Audio_to_text(audio_file=audio_path)
# df = obj.response_prompt(temperature=0.7)

# # create an instance of the VoiceParameterEvaluator class and call its assign_score method
# score_obj = VoiceParameterEvaluator()
# score_obj.assign_score(df, Path(audio_path).stem+ '.xlsx')        
        